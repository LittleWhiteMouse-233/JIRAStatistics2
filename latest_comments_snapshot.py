from models import JIRALogin, IssueList, ConcatFilter, WorksheetShell
from models import CellSetting as CSt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import time

if __name__ == '__main__':
    jira_agent = JIRALogin.used_token(r'access_token.txt')
    issue_list = IssueList(jira_agent.get_fields())
    issue_list.import_issues(jira_agent.search_by_jql_filter(ConcatFilter.EPIC_COMMENT))
    comments_table = issue_list.get_comments_status(-1)
    workbook = Workbook()
    worksheet = workbook.active
    for row_content in dataframe_to_rows(comments_table, index=False, header=False):
        worksheet.append(row_content)
    wss = WorksheetShell(worksheet)
    # 纵向合并单元格
    wss.batch_merge_cells_vertical(col_list=['A', 'B', 'C'])
    # 复制合并
    wss.copy_merge_cells_vertical('C', 'D')
    # 水平居中，垂直居中
    wss.batch_set(CSt.setting_text_alignment, col_list=['A', 'B', 'D'], horizontal='center')
    # 水平居左，垂直居中
    wss.batch_set(CSt.setting_text_alignment, col_list=['C', 'E'])
    # 列宽
    wss.batch_set_column_width({
        'A': 20,
        'B': 30,
        'C': 40,
        'D': 15,
        'E': 100,
    })
    # 自动换行
    wss.batch_set(CSt.setting_word_wrap, col_list=['B', 'C', 'E'])
    # 单元格边框
    wss.batch_set(CSt.setting_cell_border)
    # 单元格颜色
    wss.batch_set(CSt.setting_fill_color_by_re, col_list=['D'], re_pattern=re.compile(r'(<0x)([a-zA-Z0-9]{6})(>)'))
    # 单元格字体
    wss.batch_set(CSt.setting_basic_font, col_list=['D'], bold=True, color='FFFFFF')
    # 输出表格
    workbook.worksheets[0] = wss.worksheet
    filename = 'Comments snapshot at ' + time.asctime().replace(':', '-') + '.xlsx'
    # filename = 'Comments snapshot.xlsx'
    workbook.save(filename)
    print("Workbook saved as: %s." % filename)
