import pandas as pd
import numpy as np
import warnings
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from models.support.workbookProcess import WorksheetProcessor, RCActivator, HeatmapRenderer
from models.support import exceptions as exc


class ReferenceMap:
    def __init__(self, worksheet: Worksheet, origin_point_rc: tuple[int, int]):
        # origin_point_rc: 为坐标重叠域右下交点的自然行列坐标，对应值域左上起点的零点行列坐标
        self.__origin_ws = worksheet
        # 原点行坐标，原点列坐标
        self.__op_r, self.__op_c = origin_point_rc
        ws = WorksheetProcessor.copy_into(worksheet, Workbook().active)
        unmerged_sheet = WorksheetProcessor.unmerge_cells_and_fill(ws, fill=True)
        # unmerged_sheet = WorksheetProcessor(deepcopy(worksheet)).unmerge_cells_and_fill(fill=True)
        table = pd.DataFrame(unmerged_sheet.values)
        table.dropna(axis=0, how='all', inplace=True)
        table.dropna(axis=1, how='all', inplace=True)
        self.__value_map = self.__reset_rc(table.iloc[self.__op_r:, self.__op_c:])
        self.__axis_x = self.__reset_rc(table.iloc[:self.__op_r, self.__op_c:])
        self.__axis_y = self.__reset_rc(table.iloc[self.__op_r:, :self.__op_c])

    @staticmethod
    def __reset_rc(df: pd.DataFrame, row=True, col=True, row_drop=True, col_drop=True):
        if row:
            df = df.reset_index(drop=row_drop)
        if col:
            df = df.T.reset_index(drop=col_drop).T
        return df

    # 横轴层数
    @property
    def __ln_x(self):
        return self.__op_r

    # 纵轴层数
    @property
    def __ln_y(self):
        return self.__op_c

    @property
    def value_shape(self):
        return self.__value_map.shape

    @property
    def worksheet_name(self):
        return self.__origin_ws.title

    def __check_level_x(self, level_x: int):
        if not 1 <= level_x <= self.__ln_x:
            raise ValueError("The level_x(%d) is out of level-num range(1, %d)." % (level_x, self.__ln_x))

    def __check_level_y(self, level_y: int):
        if not 1 <= level_y <= self.__ln_y:
            raise ValueError("The level_y(%d) is out of level-num range(1, %d)." % (level_y, self.__ln_y))

    def __locate_coord_value(self, value: str, level: int, axis: int):
        # level: belong to N*
        if axis == 0:
            self.__check_level_x(level)
            return self.__axis_x.iloc[level - 1, :].eq(value)
        elif axis == 1:
            self.__check_level_y(level)
            return self.__axis_y.iloc[:, level - 1].eq(value)
        else:
            raise ValueError("The axis should be 0 or 1, but get %d." % axis)

    def __locate_multilayer_coord(self, coord_list: list[str], axis: int, auto_adapt=False):
        if auto_adapt:
            coord_list = self.__auto_adapt_coord_list(coord_list, axis)
        xy_ln = (self.__ln_x, self.__ln_y)
        if len(coord_list) != xy_ln[axis]:
            raise ValueError("The length of coord_list: %d is different from the level-num of axis(%d): %d."
                             % (len(coord_list), axis, xy_ln[axis]))
        bool_index = pd.Series([True] * self.value_shape[int(not bool(axis))])
        for i, coord in enumerate(coord_list):
            if coord is None:
                continue
            bool_index &= self.__locate_coord_value(coord, level=i + 1, axis=axis)
            if not bool_index.any():
                break
        return bool_index

    def __auto_adapt_coord_list(self, coord_list: list[str], axis: int):
        i = len(coord_list)
        for i in reversed(range(len(coord_list))):
            if coord_list[i] is not None:
                break
        ata_coord_list = coord_list[:i + 1]
        level_length = (self.__ln_x, self.__ln_y)[axis]
        return [None] * (level_length - len(ata_coord_list)) + ata_coord_list

    def locate_coord_cell(self, row_coordinates: list[str], col_coordinates: list[str]):
        row_index = self.__locate_multilayer_coord(row_coordinates, axis=1, auto_adapt=True)
        col_index = self.__locate_multilayer_coord(col_coordinates, axis=0, auto_adapt=True)
        located = self.__value_map.loc[row_index, col_index]
        coord = (*row_coordinates, *col_coordinates)
        if located.empty:
            raise exc.NoMatchingError(coord)
        if located.size != 1:
            raise exc.ManyMatchingError(coord)
        located_value = pd.to_numeric(located.iloc[0, 0], errors='coerce')
        if pd.isna(located_value):
            raise exc.MatchingNAError(coord)
        return located

    @staticmethod
    def cell2value(located: pd.DataFrame):
        return pd.to_numeric(located.iloc[0, 0])

    @staticmethod
    def cell2index(located: pd.DataFrame):
        return int(located.index[0]), int(located.columns[0])

    def ref_value(self, iloc_i: int, iloc_j: int):
        value = pd.to_numeric(self.__value_map.iloc[iloc_i, iloc_j], errors='coerce')
        if pd.isna(value):
            return None
        else:
            return value

    def value_array2synthesize_sheet(self, value_array: np.ndarray, heatmap=True):
        # 数组的行列数应该与参考表的数据矩阵的行列数一致
        if value_array.shape != self.value_shape:
            raise ValueError("The shape of value_array(%s) is wrong, it shall be the same as ref_map: %s)."
                             % value_array.shape, self.value_shape)
        worksheet = WorksheetProcessor.copy_into(self.__origin_ws, Workbook().active)
        # 渲染器
        renderer = HeatmapRenderer(worksheet, value_array.max(), value_array.min(),
                                   zero_color='FFFFFF') if heatmap else None
        for i in range(value_array.shape[0]):
            for j in range(value_array.shape[1]):
                value = value_array[i][j]
                ws_i = self.__ln_x + i + 1
                ws_j = self.__ln_y + j + 1
                # 热力图
                if heatmap:
                    if self.ref_value(i, j) is None:
                        renderer.colorful_value(ws_i, ws_j, value, color='D0D0D0')
                    else:
                        renderer.colorful_value(ws_i, ws_j, value)
                else:
                    worksheet.cell(ws_i, ws_j).value = value
        return worksheet

    def value_array2downmix_sheet(self, value_array: np.ndarray, level_x: int = None, level_y: int = None,
                                  heatmap=True):
        # level_x, level_y: belong to N*
        if level_x is None and level_y is None:
            warnings.warn("If you want a sheet with the same shape as the ref_map, use value_array2sheet().",
                          RuntimeWarning)
            return self.value_array2synthesize_sheet(value_array, heatmap=heatmap)
        # 数组的行列数应该与参考表的数据矩阵的行列数一致
        if value_array.shape != self.value_shape:
            raise ValueError("The shape of value_array(%s) is wrong, it shall be the same as ref_map: %s)."
                             % value_array.shape, self.value_shape)

        def downmix(array: np.ndarray, axis_series: pd.Series, axis: int):
            assert axis == 0 or axis == 1
            assert axis_series.size == array.shape[-(axis + 1)]
            concat_list = []
            dm_axis = [axis_series[0]]
            start = 0
            flag = False
            for end in range(1, axis_series.size + 1):
                if end == axis_series.size:
                    flag = True
                else:
                    coord = axis_series[end]
                    if coord != dm_axis[-1]:
                        dm_axis.append(coord)
                        flag = True
                if flag:
                    assert end > start
                    if axis == 0:
                        concat_list.append(array[:][start:end].sum(axis=1))
                    else:
                        concat_list.append(array[start:end][:].sum(axis=0))
                    start = end
                    flag = False
            assert len(concat_list) == len(dm_axis)
            if axis == 0:
                return np.vstack(concat_list).T, pd.DataFrame(dm_axis).T
            else:
                return np.vstack(concat_list), pd.DataFrame(dm_axis)

        if level_x is not None:
            self.__check_level_x(level_x)
            downmix_x, axis_x = downmix(value_array, self.__axis_x.iloc[level_x - 1, :], axis=0)
        else:
            downmix_x = value_array
            axis_x = self.__axis_x
        if level_y is not None:
            self.__check_level_y(level_y)
            downmix_xy, axis_y = downmix(downmix_x, self.__axis_y.iloc[:, level_y - 1], axis=1)
        else:
            downmix_xy = downmix_x
            axis_y = self.__axis_y
        op_r, op_c = axis_x.shape[0], axis_y.shape[1]
        worksheet = Workbook().active
        # 填充横坐标域
        if level_x is not None:
            for i in range(axis_x.shape[0]):
                for j in range(axis_x.shape[1]):
                    worksheet.cell(i + 1, op_c + j + 1).value = axis_x.iloc[i, j]
        else:
            # 复制原表的横坐标域
            WorksheetProcessor.copy_part_into(self.__origin_ws, worksheet,
                                              RCActivator.scope_int2str(1,
                                                                        self.__op_c + 1,
                                                                        self.__op_r,
                                                                        self.__op_c + self.__axis_x.shape[1]),
                                              RCActivator.point_int2str(1, op_c + 1))
        # 填充纵坐标域
        if level_y is not None:
            for i in range(axis_y.shape[0]):
                for j in range(axis_y.shape[1]):
                    worksheet.cell(op_r + i + 1, j + 1).value = axis_y.iloc[i, j]
        else:
            # 复制原表的纵坐标域
            WorksheetProcessor.copy_part_into(self.__origin_ws, worksheet,
                                              RCActivator.scope_int2str(self.__op_r + 1,
                                                                        1,
                                                                        self.__op_r + self.__axis_y.shape[0],
                                                                        self.__op_c),
                                              RCActivator.point_int2str(op_r + 1, 1))
        # 补充表头域
        if level_x is None and level_y is not None:
            # 复制压缩列对应的表头域
            WorksheetProcessor.copy_part_into(self.__origin_ws, worksheet,
                                              RCActivator.scope_int2str(1,
                                                                        level_y,
                                                                        self.__op_r,
                                                                        level_y),
                                              RCActivator.point_int2str(1, 1))
        # 填充值域
        renderer = HeatmapRenderer(worksheet, value_array.max(), value_array.min(),
                                   zero_color='FFFFFF') if heatmap else None
        for i in range(downmix_xy.shape[0]):
            for j in range(downmix_xy.shape[1]):
                ws_i, ws_j = op_r + i + 1, op_c + j + 1
                value = downmix_xy[i][j]
                if heatmap:
                    renderer.colorful_value(ws_i, ws_j, value)
                else:
                    worksheet.cell(ws_i, ws_j).value = value
        return worksheet
