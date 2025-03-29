import pandas as pd
import numpy as np
# import itertools
from enum import Enum
from typing import Callable
from openpyxl import load_workbook, Workbook
# from copy import deepcopy
from .accessAgent import JIRAOperator
from . import issueData as issueD
from . import fieldStructure as fieldS
from .component import ReferenceMap
from .support import exceptions as exc, utils
from .support.workbookProcess import WorksheetProcessor


class Workload:
    def __init__(self, worklog: fieldS.Worklog, jira_op: JIRAOperator, rate: float):
        issue = jira_op.find_issue_by(worklog.issueId)
        task, epic = jira_op.find_parents(issue)
        self.__issue: issueD.Issue = issue
        self.__task: issueD.Task = task
        self.__epic: issueD.Epic = epic
        self.__worklog = worklog
        self.__rate = rate
        # 额外信息
        self.__extend_of_issue(issue)
        self.__extend_of_task(task)
        self.__extend_of_epic(epic)

    def __extend_of_issue(self, issue: issueD.Issue):
        self.project = issue.belongingProject
        self.task_type = issue.task_type

    def __extend_of_task(self, task: issueD.Task):
        if task is not None:
            self.platform = task.platform_string
            if self.task_type is None:
                self.task_type = task.task_type
        else:
            self.platform = None

    def __extend_of_epic(self, epic: issueD.Epic):
        self.certification = epic.certification_string
        if self.platform is None:
            self.platform = epic.platform_string

    @property
    def belong_issue_key(self):
        return self.__issue.key

    @property
    def person_hour(self):
        return self.__rate * self.__worklog.timeSpentSeconds / 3600

    @property
    def person_day(self):
        return self.person_hour / 8

    def is_default(self):
        return self.__worklog.id == '-1'

    def get_worklog_info(self):
        return pd.Series([
            self.__issue.info_string,
            self.__task.info_string if self.__task else '# NoTask',
            self.__epic.info_string,
            self.project.name,
            self.__worklog.created_author.displayName,
            self.__worklog.comment,
            self.__worklog.timeSpentSeconds / 3600,
            self.__rate,
        ])


class Cell:
    def __init__(self, r1: str, r2: str, c1: str, c2: str, ref_map: ReferenceMap):
        self.__workloads: list[Workload] = []
        self.__r1 = r1
        self.__r2 = r2
        self.__c1 = c1
        self.__c2 = c2
        self.__ref_map = ref_map
        located_cell = ref_map.locate_coord_cell([self.__r1, self.__r2], [self.__c1, self.__c2])
        # (person·day)
        self.__std_time = ReferenceMap.cell2value(located_cell)
        self.__value_index = ReferenceMap.cell2index(located_cell)

    def add_workload(self, issue: issueD.Issue, jira_op: JIRAOperator, rate: float):
        if issue.worklogs:
            for worklog in issue.worklogs:
                assert worklog.issueId == issue.id
                self.__workloads.append(Workload(worklog, jira_op, rate))
        else:
            default_worklog = fieldS.Worklog(
                id='-1',
                created_author=issue.assignee if issue.assignee is not None else issue.creator,
                created_timestring=issue.resolution_timestring,
                issueId=issue.id,
                comment="# DefaultWorklog",
                started_timestring='',
                timeSpent='',
                timeSpentSeconds=self.__std_time * 8 * 3600,
                updated_author=fieldS.User.init_default(),
                updated_timestring='',
            )
            self.__workloads.append(Workload(default_worklog, jira_op, rate))

    @property
    def coord_string(self):
        return '([%s]-[%s], [%s]-[%s])' % self.coord_tuple

    @property
    def coord_tuple(self):
        return self.__r1, self.__r2, self.__c1, self.__c2

    @property
    def coord_index(self):
        return self.__value_index

    @property
    def num_worklog(self):
        return len(self.__workloads)

    @property
    def num_issues(self):
        return len(set(map(lambda x: x.belong_issue_key, self.__workloads)))

    def cumulative_workload(self, unit='day'):
        if unit == 'day':
            return sum(map(lambda x: x.person_day, self.__workloads))
        elif unit == 'hour':
            return sum(map(lambda x: x.person_hour, self.__workloads))
        else:
            raise ValueError("Unknown unit: %s." % unit)


    def standard_workload(self, unit='day'):
        if unit == 'day':
            return self.__std_time
        elif unit == 'hour':
            return self.__std_time * 8
        else:
            raise ValueError("Unknown unit: %s." % unit)

    def refer_from(self, ref_map: ReferenceMap):
        if ref_map is self.__ref_map or ref_map.worksheet_name == self.__ref_map.worksheet_name:
            return True
        else:
            return False

    def get_worklog_table(self):
        concat_list = []
        for workload in self.__workloads:
            concat_list.append(workload.get_worklog_info())
        table = pd.concat(concat_list, axis=1).T
        table = utils.concat_single_value(table,
                                          left=[self.coord_string],
                                          repeat=True,
                                          columns=['coordinate', 'issue', 'task', 'epic', 'project',
                                                   'creator', 'comment', 'time(hour)', 'rate'])
        return table


class Matrix:
    class LoadResult(Enum):
        SKIP = 1, 'Skip'
        SUCCESS = 0, 'Success'
        WRONG = -1, 'Wrong'

        def __init__(self, result_id: int, result_name: str):
            self.res_id = result_id
            self.res_name = result_name

    class __MetaData:
        def __init__(self, issue: issueD.Issue | issueD.TaskLike):
            self.__issue = issue
            self.__load_result = None
            self.__load_detail = None
            self.ref_class = type(issue)
            self.std_time = -1
            if issue.worklogs:
                self.worklog = (sum(map(lambda x:x.timeSpentSeconds, issue.worklogs)))/3600/8
            else:
                self.worklog = -1

        @property
        def issue(self):
            return self.__issue

        @property
        def load_result(self):
            return self.__load_result

        @property
        def load_detail(self):
            return self.__load_detail

        def skip(self, detail: str = None):
            self.__load_result = Matrix.LoadResult.SKIP
            self.__load_detail = detail

        def success(self, detail: str = None):
            self.__load_result = Matrix.LoadResult.SUCCESS
            self.__load_detail = detail

        def wrong(self, detail: str = None):
            self.__load_result = Matrix.LoadResult.WRONG
            self.__load_detail = detail

    def __init__(self, issues: issueD.IssueList, jira_op: JIRAOperator, ref_filename: str):
        jira_op.add_cache(issues)
        self.__jira_op = jira_op
        ref_xlsx = load_workbook(ref_filename)
        self.__ref_test = ReferenceMap(ref_xlsx.worksheets[0], (2, 3))
        self.__ref_manage = ReferenceMap(ref_xlsx.worksheets[1], (2, 3))
        self.__meta_datas: list[Matrix.__MetaData] = []
        self.__cells: list[Cell] = []
        for issue in issues:
            self.__meta_datas.append(self.__MetaData(issue))
        for metadata in self.__meta_datas:
            issue = metadata.issue
            if type(issue) is issueD.Subtask:
                try:
                    p1, p2 = jira_op.find_parents(issue)
                except exc.GetIssueFailedError as e:
                    metadata.wrong("%s(TaskLike.parent_key=%s)." % (e, metadata.issue.parent_key))
                    continue
                if p1 is None:
                    metadata.wrong('该子任务的父任务为 Epic！！！')
                    continue
            # 非类任务型
            if not issubclass(type(issue), issueD.TaskLike):
                metadata.skip("Is not subclass of TaskLike: %s." % issue.issueType.name)
                continue
            index = issues.self_search_by(issue.parent_key, return_index=True)
            # 排除存在子任务的父任务
            if index is not None:
                md_p = self.__meta_datas[index]
                md_p.skip("This issue is parent of: %s." % metadata.issue.key)
        self.load_workload_into_cell()

    # @staticmethod
    # def __coordinate_grouping(coord: tuple):
    #     group_list = []
    #     for x in coord:
    #         if type(x) is dict:
    #             group_list.append(list(x.values()))
    #         else:
    #             group_list.append([x])
    #     coord_group = itertools.product(*group_list)
    #     return list(coord_group)
    #
    # @staticmethod
    # def __coordinate_generator(coord: tuple):
    #     def ranging(rc: str | tuple[str] | list[str]):
    #         if type(rc) is tuple or type(rc) is list:
    #             return rc
    #         return (rc,)
    #
    #     r1, r2, c1, c2 = coord
    #     for i1 in ranging(r1):
    #         for i2 in ranging(r2):
    #             for j1 in ranging(c1):
    #                 for j2 in ranging(c2):
    #                     i1: str
    #                     i2: str
    #                     j1: str
    #                     j2: str
    #                     yield i1, i2, j1, j2

    # def __analyse_coordinate(self, metadata: __MetaData):
    #     task_like = metadata.issue
    #     # 构造事务链
    #     task, epic = self.__jira_op.find_parents(task_like)
    #     if type(task_like) is issueD.Subtask and issubclass(type(task), issueD.TaskLike):
    #         metadata.ref_class = type(task)
    #     # 生成坐标（集合）
    #     coord = task_like.generate_coordinate(epic, task=task)
    #     # 区分主副坐标集（例如“认证项”字段）
    #     coord_group = self.__coordinate_grouping(coord)
    #     # 拆分坐标集（例如“模块”字段），构建坐标生成器
    #     coord_gens = []
    #     for group in coord_group:
    #         coord_gens.append(self.__coordinate_generator(group))
    #     return coord_gens

    def __analyse_coordinate(self, metadata: __MetaData):
        task_like = metadata.issue
        # 构造事务链
        task, epic = self.__jira_op.find_parents(task_like)
        # 子任务继承上级事务类型
        if type(task_like) is issueD.Subtask and issubclass(type(task), issueD.TaskLike):
            metadata.ref_class = type(task)
        # 生成坐标集合（分组）
        coord_cache = task_like.generate_coordinate(epic, task=task)
        return coord_cache

    def load_workload_into_cell(self):
        print("Loading workload into cell ...")
        for metadata in self.__meta_datas:
            if metadata.load_result is not None:
                continue
            try:
                # coord_gens = self.__analyse_coordinate(metadata)
                coord_cache = self.__analyse_coordinate(metadata)
            # 构造事务链失败
            except exc.GetIssueFailedError as e:
                metadata.wrong("%s(TaskLike.parent_key=%s)." % (e, metadata.issue.parent_key))
                continue
            # 生成坐标失败
            except exc.CoordinateError as e:
                metadata.wrong(str(e))
                continue
            # cell_list = []
            # wrong_flag = False
            # for coord_generator in coord_gens:
            #     try:
            #         for coord in coord_generator:
            #             cell = self.__find_cell_or_create(coord, metadata.ref_class)
            #             cell_list.append(cell)
            #         break
            #     # 匹配坐标（集合）失败
            #     except exc.MisMatchingError as e:
            #         if coord_generator is coord_gens[-1]:
            #             metadata.wrong(str(e))
            #             wrong_flag = True
            #             break
            #         else:
            #             continue
            # # 分组的所有坐标（集合）都匹配失败
            # if wrong_flag:
            #     continue
            cell_list = []
            load_results = dict()
            for coord_set in coord_cache.grouping():
                try:
                    for coord in coord_set.generator:
                        cell = self.__find_cell_or_create((*coord[0], *coord[1]), metadata.ref_class)
                        cell_list.append(cell)
                # 匹配坐标集失败
                except exc.MisMatchingError as e:
                    cell_list = []
                    load_results[coord_set.cs_label] = str(e)
                    continue
                # 任意一组坐标集匹配成功
                else:
                    break
            # 所有坐标集都匹配失败
            if not cell_list:
                metadata.wrong(str(load_results))
                continue
            for cell in cell_list:
                cell.add_workload(metadata.issue, self.__jira_op,
                                  cell.standard_workload() / sum(map(lambda x: x.standard_workload(), cell_list)))
            metadata.success("Coordinate(s): " + ' & '.join([cell.coord_string for cell in cell_list]))
            metadata.std_time = sum(map(lambda x: x.standard_workload(), cell_list))
            if metadata.worklog == -1:
                metadata.worklog = sum(map(lambda x: x.standard_workload(), cell_list))
        assert len(self.__meta_datas) == sum(self.__num_of(res) for res in self.LoadResult)
        print("Loading issue completed.\n")

    def __find_cell_or_create(self, ref_coord: tuple, ref_class: type):
        flag = True
        for cell in self.__cells:
            coord = cell.coord_tuple
            for i in range(len(coord)):
                if coord[i] != ref_coord[i]:
                    flag = False
                    break
            if flag:
                return cell
            flag = True
        # 已有的 Cell 没有坐标能对应上，新建 Cell
        if ref_class is issueD.TestTask:
            ref_map = self.__ref_test
        elif ref_class is issueD.ManageTask:
            ref_map = self.__ref_manage
        else:
            ref_map = None
        new_cell = Cell(*ref_coord, ref_map=ref_map)
        self.__cells.append(new_cell)
        return new_cell

    def export_worklog_table(self):
        print("Exporting worklog table ...")
        concat_list = []
        for cell in self.__cells:
            if cell.num_worklog == 0:
                continue
            concat_list.append(cell.get_worklog_table())
        table = pd.concat(concat_list).reset_index(drop=True)
        table.sort_values(by=list(table.columns[[0, 1, 4]]), inplace=True)
        print("Exporting worklog table completed.\n")
        return table

    def __num_of(self, res: LoadResult):
        return sum(map(lambda x: x.load_result is res, self.__meta_datas))

    def meta_data_loading_report(self, show_detail=True):
        def temp_func_worklog(md:Matrix.__MetaData):

            return
        print("Loading report: ")
        report_list = []
        if show_detail:
            for res in self.LoadResult:
                print("\t%s(%s): " % (res.res_name, self.__num_of(res)))
                for metadata in self.__meta_datas:
                    if metadata.load_result is res:
                        report_list.append([
                            metadata.issue.key,
                            metadata.issue.issueType.name,
                            metadata.issue.summary,
                            metadata.issue.creator.displayName,
                            metadata.issue.assignee.displayName if metadata.issue.assignee is not None else '未指定',
                            str(metadata.load_result),
                            metadata.load_detail,
                            metadata.std_time,
                            metadata.worklog,
                        ])
                        print(utils.specific_length_string(metadata.issue.info_string), metadata.load_detail)
        print('Total: %d' % len(self.__meta_datas), end=', ')
        print(', '.join(map(lambda x: '%s: %d' % (x.res_name, self.__num_of(x)), self.LoadResult)) + '\n')
        return pd.DataFrame(report_list, columns=[
            'issue_key',
            'issue_type',
            'summary',
            'creator',
            'assignee',
            'result',
            'detail',
            '标准工时（人·天），如无有效值则为-1',
            '结算工时（人·天），如无有效值则为-1',
        ])

    def __build_matrix(self, ref_map: ReferenceMap, cell_property: Callable[[Cell], int | float]):
        array = np.zeros(ref_map.value_shape)
        for cell in self.__cells:
            if not cell.refer_from(ref_map):
                continue
            row_index, col_index = cell.coord_index
            array[row_index][col_index] = cell_property(cell)
        return array

    def __synthesize_sheet(self, ref_map: ReferenceMap, cell_property: Callable[[Cell], int | float], head='{}'):
        value_array = self.__build_matrix(ref_map, cell_property)
        worksheet = ref_map.value_array2synthesize_sheet(value_array)
        worksheet.cell(1, 1).value = head.format(ref_map.worksheet_name)
        return worksheet

    def __downmix_sheet(self, ref_map: ReferenceMap, downmix_x: int | None, downmix_y: int | None,
                        cell_property: Callable[[Cell], int | float], head='{}'):
        value_array = self.__build_matrix(ref_map, cell_property)
        worksheet = ref_map.value_array2downmix_sheet(value_array, downmix_x, downmix_y)
        worksheet.cell(1, 1).value = (head.format(ref_map.worksheet_name)
                                      + ' downmix by (%s, %s)' % (downmix_x, downmix_y))
        return worksheet

    def export_matrix_workbook(self):
        unit = 'day'

        def cell_count(cell: Cell):
            return cell.num_issues

        def cell_workload(cell: Cell):
            return cell.cumulative_workload(unit=unit)

        count_head = r"Count of {}"
        workload_head = r"Cumulative workload of {}(person·%s)" % unit
        print("Exporting matrix workbook ...")
        workbook = Workbook()
        worksheet = workbook.active
        # 测试计数
        print("Building count matrix and synthesizing with ref_test style ...")
        test_count = self.__synthesize_sheet(self.__ref_test, cell_count, count_head)
        WorksheetProcessor.copy_into(test_count, worksheet)
        worksheet.title = 'Count of Test'
        # 测试计数压缩
        print("Building DOWNMIX count matrix base on ref_test ...")
        test_count_dm = self.__downmix_sheet(self.__ref_test, None, 1, cell_count, count_head)
        WorksheetProcessor.copy_into(test_count_dm, workbook.create_sheet('Count of Test(DOWNMIX)'))
        # 测试计时
        print("Building workload matrix and synthesizing with ref_test style ...")
        test_workload = self.__synthesize_sheet(self.__ref_test, cell_workload, workload_head)
        WorksheetProcessor.copy_into(test_workload, workbook.create_sheet('Time of Test'))
        # 测试计时压缩
        print("Building DOWNMIX workload matrix base on ref_test ...")
        test_workload_dm = self.__downmix_sheet(self.__ref_test, None, 1, cell_workload, workload_head)
        WorksheetProcessor.copy_into(test_workload_dm, workbook.create_sheet('Time of Test(DOWNMIX)'))
        # 管理计数
        print("Building count matrix and synthesizing with ref_manage style ...")
        manage_count = self.__synthesize_sheet(self.__ref_manage, cell_count, count_head)
        WorksheetProcessor.copy_into(manage_count, workbook.create_sheet('Count of Manage'))
        # 管理计数压缩
        pass
        # 管理计时
        print("Building workload matrix and synthesizing with ref_manage style ...")
        manage_workload = self.__synthesize_sheet(self.__ref_manage, cell_workload, workload_head)
        WorksheetProcessor.copy_into(manage_workload, workbook.create_sheet('Time of Manage'))
        print("Exporting matrix workbook completed\n")
        # 管理计时压缩
        pass
        return workbook

    def workload_analyzer(self):
        pass
