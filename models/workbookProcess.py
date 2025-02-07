import re
from typing import Callable, Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, PatternFill
from matplotlib.colors import LinearSegmentedColormap, Colormap
from copy import copy


class RCActivator:
    @staticmethod
    def alpha2num(alpha: str):
        if not alpha.isalpha():
            raise ValueError("The parameter \'alpha\' string has non-alphabetic characters: " + str(alpha))
        num = 0
        for i, c in enumerate(alpha[::-1]):
            num += (ord(c.upper()) - ord('A') + 1) * pow(26, i)
        return num

    @staticmethod
    def num2alpha(num: int):
        if num <= 0:
            raise ValueError("The parameter \'decimal_num\'(" + str(num) + ") cannot be less than 0.")
        alpha = ''
        while True:
            num, mod = divmod(num, 26)
            alpha = chr(mod + ord('A') - 1) + alpha
            if num == 0:
                break
        return alpha

    @classmethod
    def point_str2int(cls, point_str: str):
        return (int(re.search(r'\d+', point_str).group()),
                cls.alpha2num(re.search(r'[A-Z]+', point_str).group()))

    @classmethod
    def point_int2str(cls, row: int, col: int):
        if row <= 0:
            raise ValueError("The value of row cannot be less than 0.")
        if col <= 0:
            raise ValueError("The value of col cannot bu less than 0.")
        return cls.num2alpha(col) + str(row)

    # 数字化列名称 & 超限检查
    @classmethod
    def activate_col(cls, worksheet: Worksheet, col: int | str):
        if type(col) is str:
            act_col = cls.alpha2num(col)
        else:
            act_col: int = col
        if type(act_col) is not int:
            raise ValueError("Invalid type of act_col: %s, excepted int or str" % type(act_col))
        if not 1 <= act_col <= worksheet.max_column:
            raise ValueError("The act_col(%d) is out of range(%d:%d)" % (act_col, 1, worksheet.max_column))
        return act_col

    # 数字化列名称列表 & 超限检查
    @classmethod
    def activate_col_list(cls, worksheet: Worksheet, col_list: list[int | str]):
        return sorted(list(set(map(lambda col: cls.activate_col(worksheet, col), col_list))))

    # 字符化行列区域数字
    @classmethod
    def scope_int2str(cls, start_row: int, start_col: int, end_row: int, end_col: int):
        return cls.point_int2str(start_row, start_col) + ':' + cls.point_int2str(end_row, end_col)

    # 数字化行列区域字符串
    @classmethod
    def scope_str2int(cls, scope_str: str):
        if ':' not in scope_str:
            raise ValueError("Invalid scope string: " + str(scope_str))
        begin, end = scope_str.upper().split(':')
        begin_row, begin_col = cls.point_str2int(begin)
        end_row, end_col = cls.point_str2int(end)

        def sort_xy(x: int, y: int):
            if x <= y:
                return x, y
            else:
                return y, x

        row_scope = sort_xy(begin_row, end_row)
        col_scope = sort_xy(begin_col, end_col)
        return row_scope, col_scope

    # 数字化行列区域字符串 & 超限检查
    @classmethod
    def activate_scope(cls, worksheet: Worksheet, scope_str: str):
        max_row, max_col = worksheet.max_row, worksheet.max_column
        if not scope_str:
            return (1, max_row), (1, max_col)
        row_scope, col_scope = cls.scope_str2int(scope_str)
        if not 1 <= row_scope[0] <= row_scope[1] <= max_row:
            raise ValueError("The row(%s) is out of range(%d:%d)" % (str(row_scope), 1, max_row))
        if not 1 <= col_scope[0] <= col_scope[1] <= max_col:
            raise ValueError("The col(%s) is out of range(%d:%d)" % (str(col_scope), 1, max_col))
        return row_scope, col_scope

    # 区域内坐标生成器
    @staticmethod
    def scope_coord_generator(begin_row: int, end_row: int, begin_col: int, end_col: int):
        for i in range(begin_row, end_row + 1):
            for j in range(begin_col, end_col + 1):
                yield i, j

    # 单边（行或列）端点区域转序列范围
    @staticmethod
    def rc_scope2range(start: int, end: int):
        if start <= end:
            return range(start, end + 1)
        else:
            return range(start, end - 1, -1)


class WorksheetProcessor:
    # 解除合并单元格并填充
    @staticmethod
    def unmerge_cells_and_fill(ws: Worksheet, fill=True):
        scope_list = []
        for merge_area in ws.merged_cells:
            scope_list.append((merge_area.min_row, merge_area.min_col, merge_area.max_row, merge_area.max_col))
        for min_row, min_col, max_row, max_col in scope_list:
            ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            if fill:
                for i in range(min_row, max_row + 1):
                    for j in range(min_col, max_col + 1):
                        ws.cell(i, j).value = ws.cell(min_row, min_col).value
        return ws

    # 向 target_ws 复制 source_ws 的全部内容及格式(refer to Workbook.copy_worksheet())
    @staticmethod
    def _copy_into(source_ws: Worksheet, target_ws: Worksheet):
        if type(target_ws) is not Worksheet:
            raise TypeError("The type of target should be Worksheet.")
        if target_ws is source_ws:
            raise ValueError("The target should not be itself.")
        target_wb: Workbook = target_ws.parent
        if target_wb is not source_ws.parent:
            source_wb: Workbook = source_ws.parent
            target_wb._fonts = getattr(source_wb, '_fonts')
            target_wb._fills = getattr(source_wb, '_fills')
            target_wb._borders = getattr(source_wb, '_borders')
            target_wb._alignments = getattr(source_wb, '_alignments')
        for i in range(1, source_ws.max_row + 1):
            for j in range(1, source_ws.max_column + 1):
                from_cell = source_ws.cell(i, j)
                to_cell = target_ws.cell(i, j)
                # target_cell._value = source_cell._value
                to_cell.value = from_cell.value
                to_cell.data_type = from_cell.data_type
                if from_cell.has_style:
                    # target_cell._style = copy(source_cell._style)
                    to_cell._style = copy(getattr(from_cell, '_style'))
                if from_cell.hyperlink:
                    # target_cell._hyperlink = copy(source_cell.hyperlink)
                    to_cell.hyperlink = copy(from_cell.hyperlink)
                if from_cell.comment:
                    to_cell.comment = copy(from_cell.comment)
        for attr in ('row_dimensions', 'column_dimensions'):
            from_dim = getattr(source_ws, attr)
            to_dim = getattr(target_ws, attr)
            for key, dim in from_dim.items():
                to_dim[key] = copy(dim)
                to_dim[key].worksheet = target_ws
        target_ws.sheet_format = copy(source_ws.sheet_format)
        target_ws.sheet_properties = copy(source_ws.sheet_properties)
        target_ws.merged_cells = copy(source_ws.merged_cells)
        target_ws.page_margins = copy(source_ws.page_margins)
        target_ws.page_setup = copy(source_ws.page_setup)
        target_ws.print_options = copy(source_ws.print_options)
        # target_ws._images = self.__ws._images
        return target_ws

    # 向 target_ws 复制 source_ws 的全部内容及格式
    @staticmethod
    def copy_into(source_ws: Worksheet, target_ws: Worksheet):
        if type(target_ws) is not Worksheet:
            raise TypeError("The type of target should be Worksheet.")
        if target_ws is source_ws:
            raise ValueError("The target should not be itself.")
        for i in range(1, source_ws.max_row + 1):
            for j in range(1, source_ws.max_column + 1):
                src_cell = source_ws.cell(i, j)
                tar_cell = target_ws.cell(i, j)
                tar_cell.value = src_cell.value
                if src_cell.has_style:
                    tar_cell.alignment = copy(src_cell.alignment)
                    tar_cell.border = copy(src_cell.border)
                    tar_cell.fill = copy(src_cell.fill)
                    tar_cell.font = copy(src_cell.font)
        for key, dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[key].height = dim.height
        for key, dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[key].width = dim.width
        target_ws.merged_cells = copy(source_ws.merged_cells)
        return target_ws

    # 向 target_ws 复制 source_ws 的部分内容及格式
    @staticmethod
    def copy_part_into(source_ws: Worksheet, target_ws: Worksheet, from_scope='', to_op=''):
        if type(target_ws) is not Worksheet:
            raise TypeError("The type of target should be Worksheet.")
        if target_ws is source_ws:
            raise ValueError("The target should not be itself.")
        from_row_scope, from_col_scope = RCActivator.activate_scope(source_ws, from_scope)
        from_row_range = range(from_row_scope[0], from_row_scope[1] + 1)
        from_col_range = range(from_col_scope[0], from_col_scope[1] + 1)
        if to_op:
            to_op_r, to_op_c = RCActivator.point_str2int(to_op)
            offset_r, offset_c = to_op_r - from_row_scope[0], to_op_c - from_col_scope[0]
        else:
            offset_r, offset_c = 0, 0
        for from_row in from_row_range:
            for from_col in from_col_range:
                src_cell = source_ws.cell(from_row, from_col)
                tar_cell = target_ws.cell(from_row + offset_r, from_col + offset_c)
                tar_cell.value = src_cell.value
                if src_cell.has_style:
                    tar_cell.alignment = copy(src_cell.alignment)
                    tar_cell.border = copy(src_cell.border)
                    tar_cell.fill = copy(src_cell.fill)
                    tar_cell.font = copy(src_cell.font)
        for key, dim in source_ws.row_dimensions.items():
            if key not in from_row_range:
                continue
            target_ws.row_dimensions[key + offset_r].height = dim.height
        for key, dim in source_ws.column_dimensions.items():
            key_n = RCActivator.alpha2num(key)
            if key_n not in from_col_range:
                continue
            target_ws.column_dimensions[RCActivator.num2alpha(key_n + offset_c)].width = dim.width
        for merge_area in source_ws.merged_cells:
            min_row, min_col, max_row, max_col = (
                merge_area.min_row, merge_area.min_col, merge_area.max_row, merge_area.max_col)
            if (min_row in from_row_range and max_row in from_row_range and
                    min_col in from_col_range and max_col in from_col_range):
                target_ws.merge_cells(start_row=min_row + offset_r,
                                      start_column=min_col + offset_c,
                                      end_row=max_row + offset_r,
                                      end_column=max_col + offset_c)
        return target_ws


class CellSetting:
    ALIGNMENT_HORIZONTAL = ['general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous', 'distributed']
    ALIGNMENT_VERTICAL = ['top', 'center', 'bottom', 'justify', 'distributed']
    BORDER_STYLE = ["dashDot", "dashDotDot", "dashed", "dotted", "double", "hair", "medium", "mediumDashDot",
                    "mediumDashDotDot", "mediumDashed", "slantDashDot", "thick", "thin", "none"]
    FILL_TYPE = ["solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal", "darkTrellis", "darkUp", "darkVertical",
                 "gray0625", "gray125", "lightDown", "lightGray", "lightGrid", "lightHorizontal", "lightTrellis",
                 "lightUp", "lightVertical", "mediumGray", "none"]

    # 设置单元格文本对齐
    @classmethod
    def setting_text_alignment(cls, ws: Worksheet, row: int, col: int, horizontal='left', vertical='center'):
        if horizontal not in cls.ALIGNMENT_HORIZONTAL:
            raise ValueError("Invalid horizontal: %s, excepted input: %s." % (horizontal, cls.ALIGNMENT_HORIZONTAL))
        if vertical not in cls.ALIGNMENT_VERTICAL:
            raise ValueError("Invalid vertical: %s, excepted input: %s." % (vertical, cls.ALIGNMENT_VERTICAL))
        align = ws.cell(row, col).alignment
        if align.horizontal != horizontal or align.vertical != vertical:
            align_new = copy(align)
            if align.horizontal != horizontal:
                align_new.horizontal = horizontal
            if align.vertical != vertical:
                align_new.vertical = vertical
            ws.cell(row, col).alignment = align_new

    # 设置单元格文本自动换行
    @staticmethod
    def setting_word_wrap(ws: Worksheet, row: int, col: int):
        align = ws.cell(row, col).alignment
        if not align.wrapText:
            align_new = copy(align)
            align_new.wrapText = True
            ws.cell(row, col).alignment = align_new

    # 设置单元格边框
    @classmethod
    def setting_cell_border(cls, ws: Worksheet, row: int, col: int, border_style='thin'):
        if border_style not in cls.BORDER_STYLE:
            raise ValueError("Invalid border_style: %s, excepted input: %s" % (border_style, cls.BORDER_STYLE))
        ws.cell(row, col).border = Border(left=Side(border_style=border_style, color='FF000000'),
                                          right=Side(border_style=border_style, color='FF000000'),
                                          top=Side(border_style=border_style, color='FF000000'),
                                          bottom=Side(border_style=border_style, color='FF000000'))

    # 设置单元格颜色
    @classmethod
    def setting_fill_color(cls, ws: Worksheet, row: int, col: int, color: str, fill_type='solid'):
        if fill_type not in cls.FILL_TYPE:
            raise ValueError("Invalid fill_type: %s, excepted input: %s" % (fill_type, cls.FILL_TYPE))
        type_index = cls.FILL_TYPE.index(fill_type)
        ws.cell(row, col).fill = PatternFill(patternType=cls.FILL_TYPE[type_index], fgColor=color.upper())

    # 设置单元格颜色 by re
    @classmethod
    def setting_fill_color_by_re(cls, ws: Worksheet, row: int, col: int, re_pattern: re.Pattern, fill_type='solid'):
        content = ws.cell(row, col).value
        search_result = re_pattern.search(content)
        if not search_result:
            return
        color = search_result.groups()[1]
        ws.cell(row, col).value = re_pattern.sub('', content)
        cls.setting_fill_color(ws, row, col, color, fill_type)

    # 设置单元格颜色 by picker
    @classmethod
    def setting_fill_color_by_picker(cls, ws: Worksheet, row: int, col: int, color_picker: Callable[[Any], str],
                                     fill_type='solid'):
        content = ws.cell(row, col).value
        # try:
        #     num = float(content)
        # except ValueError:
        #     return
        color = color_picker(content)
        if color:
            cls.setting_fill_color(ws, row, col, color, fill_type)

    # 设置单元格文本字体
    @staticmethod
    def setting_basic_font(ws: Worksheet, row: int, col: int, name: str = None, size: int = None, bold: bool = None,
                           color: str = None, italic: bool = None, strike: bool = None):
        font_new = copy(ws.cell(row, col).font)
        if name:
            font_new.name = name
        if size:
            font_new.size = size
        if bold:
            font_new.bold = bold
        if color:
            font_new.color = color
        if italic:
            font_new.italic = italic
        if strike:
            font_new.strike = strike
        ws.cell(row, col).font = font_new


class WorksheetShell:
    __slots__ = '__ws'

    def __init__(self, worksheet: Worksheet):
        self.__ws = worksheet

    @property
    def worksheet(self):
        return self.__ws

    @property
    def max_row(self):
        return self.__ws.max_row

    @property
    def max_col(self):
        return self.__ws.max_column

    def _activate_col(self, col: int | str):
        return RCActivator.activate_col(self.__ws, col)

    def _activate_col_list(self, col_list: list[int | str]):
        return RCActivator.activate_col_list(self.__ws, col_list)

    def _activate_scope(self, scope_str: str):
        return RCActivator.activate_scope(self.__ws, scope_str)

    def __merge_cells_vertical(self, row_begin: int, row_end: int, col_begin: int, col_end: int, mode: str):
        excepted_mode = ['nan', 'same', 'all']
        if mode not in excepted_mode:
            raise ValueError("Merge mode should be: " + str(excepted_mode) + ", but given: " + str(mode))
        for col in range(col_begin, col_end + 1):
            start = row_begin
            end = row_begin
            last_str = ''
            for row in range(row_begin, row_end + 2):
                if row <= row_end:
                    cell_content = self.__ws.cell(row, col).value
                else:
                    cell_content = ''
                if (row > row_end
                        or (mode == 'nan' and type(cell_content) is str)
                        or (mode == 'same' and cell_content != last_str)
                        or (mode == 'all' and type(cell_content) is str and cell_content != last_str)):
                    if row_begin <= start < end <= row_end:
                        self.__ws.merge_cells(':'.join([RCActivator.point_int2str(start, col),
                                                        RCActivator.point_int2str(end, col)]))
                    start = row
                    if mode in ['same', 'all']:
                        last_str = cell_content
                end = row

    # 批量设置纵向单元格合并
    def batch_merge_cells_vertical(self, *, scope='', col_list: list = None, mode='all'):
        if col_list:
            act_col_list = self._activate_col_list(col_list)
            for col in act_col_list:
                self.__merge_cells_vertical(1, self.max_row, col, col, mode=mode)
        else:
            row_scope, col_scope = self._activate_scope(scope)
            self.__merge_cells_vertical(*row_scope, *col_scope, mode=mode)

    # 复制纵向单元格合并
    def copy_merge_cells_vertical(self, refer_col: int | str, target_col: int | str | list[int | str]):
        act_refer_col = self._activate_col(refer_col)
        if type(target_col) is list:
            act_target_col = self._activate_col_list(target_col)
        else:
            act_target_col = [self._activate_col(target_col)]
        merge_list = []
        for merge_area in self.__ws.merged_cells:
            if merge_area.min_col == merge_area.max_col == act_refer_col:
                merge_list.append((merge_area.min_row, merge_area.max_row))
        for col in act_target_col:
            for begin, end in merge_list:
                self.__ws.merge_cells(':'.join([RCActivator.point_int2str(begin, col),
                                                RCActivator.point_int2str(end, col)]))

    # 批量设置列宽
    def batch_set_column_width(self, width_dict: dict[int | str, int]):
        for col, width in width_dict.items():
            if not isinstance(width, (int, float)):
                raise ValueError("Invalid width type: %s." % type(width))
            if width <= 0:
                raise ValueError("The column width should be positive, but given: %s" % width)
            act_col = self._activate_col(col)
            self.__ws.column_dimensions[RCActivator.num2alpha(act_col)].width = width
        return self.__ws

    # 批量设置单元格
    def batch_set(self, func: Callable[[Worksheet, int, int], None] | Callable[[Worksheet, int, int, ...], None],
                  scope='', col_list: list = None, **kwargs):
        if col_list:
            col_range = self._activate_col_list(col_list)
            row_range = range(1, self.max_row + 1)
        else:
            row_scope, col_scope = self._activate_scope(scope)
            row_range = range(row_scope[0], row_scope[1] + 1)
            col_range = range(col_scope[0], col_scope[1] + 1)
        for i in row_range:
            for j in col_range:
                func(self.__ws, i, j, **kwargs)


class HeatmapRenderer(WorksheetShell):
    DEFAULT_COLORMAP = LinearSegmentedColormap.from_list("custom", [(0, 1, 0), (1, 1, 0), (1, 0, 0)])
    __slots__ = '__max', '__min', '__cmap', '__0c', '__0m'

    def __init__(self, worksheet: Worksheet, max_value: int | float, min_value: int | float,
                 color_map: Colormap = None, zero_color='D0D0D0', zero_mask=0):
        super().__init__(worksheet)
        self.__max = max_value
        self.__min = min_value
        if color_map is not None:
            self.__cmap = color_map
        else:
            self.__cmap = self.DEFAULT_COLORMAP
        self.__0c = zero_color
        self.__0m = zero_mask

    def __color_picker(self, scale: float):
        r, g, b, _ = self.__cmap(scale, bytes=True)
        return '%06x' % ((int(r) << 16) + (int(g) << 8) + int(b))

    def __linear_color(self, value: Any):
        if value == 0 and self.__0c:
            return self.__0c
        return self.__color_picker((value - self.__min) / (self.__max - self.__min))

    def __mask_value(self, value: Any):
        if value == 0:
            return self.__0m
        return value

    def colorful_value(self, row: int, col: int, value: Any):
        self.worksheet.cell(row, col).value = self.__mask_value(value)
        CellSetting.setting_fill_color(self.worksheet, row, col, self.__linear_color(value))
